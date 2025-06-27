import os
import openpyxl
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import Q, Sum
from django.contrib import messages
from .models import Product, Category, Manufacturer, Trash, TrashElement, Order, OrderItem
from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from . import specialty_search
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from .forms import CheckoutForm
  
def index(request):
    return HttpResponse("Приложение для управления работает!")

def author(request):
    return HttpResponse("<div><h1>Лабу сделал:</h1><h2>Имя: Владислав</h2><h2>Фамилия: Цветков</h2><h2>Отчество: Артурович</h2><h2>Группа: 82ТП</h2></div>")

def shop(request):
    return HttpResponse('<div><h1>Тема лабы:</h1><h2 style="width: 50%">Магазин наборов для рисования граффити (баллончики, трафареты).</h2></div>')

def specialty_byid(request, id):
    return HttpResponse(specialty_search.spSearch(by_id=id))

def specialty(request):
    return HttpResponse(specialty_search.spSearch())

def product_list(request):
    products = Product.objects.all()

    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)

    manufacturer_id = request.GET.get('manufacturer')
    if manufacturer_id:
        products = products.filter(manufacturer_id=manufacturer_id)

    search_query = request.GET.get('search')
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()

    context = {
        'products': products,
        'categories': categories,
        'manufacturers': manufacturers,
        'selected_category': int(category_id) if category_id else None,
        'selected_manufacturer': int(manufacturer_id) if manufacturer_id else None,
        'search_query': search_query or '',
    }

    return render(request, 'shop/products/product_list.html', context)


def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    related_products = Product.objects.filter(
        category=product.category
    ).exclude(id=product.id)[:4]

    context = {
        'product': product,
        'related_products': related_products
    }
    return render(request, 'shop/products/product_detail.html', context)


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    trash, created = Trash.objects.get_or_create(user=request.user)

    cart_item, item_created = TrashElement.objects.get_or_create(
        trash=trash,
        product=product,
        defaults={'quantity': 1}
    )

    if not item_created:
        cart_item.quantity += 1
        cart_item.save()

    messages.success(request, f'Товар "{product.name}" добавлен в корзину')
    return redirect('cart_view')


@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(TrashElement, id=item_id, trash__user=request.user)

    if request.method == 'POST':
        try:
            new_quantity = int(request.POST.get('quantity'))
            if new_quantity < 1:
                messages.error(request, "Количество не может быть меньше 1")
            elif new_quantity > cart_item.product.quantity:
                messages.error(request,
                               f"Недостаточно товара на складе. Доступно: {cart_item.product.quantity}")
            else:
                cart_item.quantity = new_quantity
                cart_item.save()
                messages.success(request, "Количество обновлено")
        except (ValueError, TypeError):
            messages.error(request, "Некорректное количество")

    return redirect('cart_view')


@login_required
def checkout(request):
    trash = get_object_or_404(Trash, user=request.user)
    cart_items = trash.items.select_related('product')

    if not cart_items.exists():
        return redirect('cart_view')

    total = sum(item.cost() for item in cart_items)

    if request.method == 'POST':
        form = CheckoutForm(request.POST)
        if form.is_valid():
            order = Order.objects.create(
                user=request.user,
                shipping_address=form.cleaned_data['shipping_address'],
                total_price=total
            )

            for item in cart_items:
                OrderItem.objects.create(
                    order=order,
                    product=item.product,
                    quantity=item.quantity,
                    price=item.product.price
                )

                item.product.quantity -= item.quantity
                item.product.save()

            excel_file = generate_order_excel(order)

            send_order_email(request.user, order, excel_file)

            trash.items.all().delete()

            return redirect('order_success', order_id=order.id)
    else:
        form = CheckoutForm()

    context = {
        'form': form,
        'cart_items': cart_items,
        'total': total
    }
    return render(request, 'shop/checkout.html', context)


def generate_order_excel(order):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказ"

    headers = ["№", "Товар", "Количество", "Цена за ед.", "Сумма"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = openpyxl.styles.Font(bold=True)

    row_num = 2
    for item in order.items.all():
        ws[f'A{row_num}'] = row_num - 1
        ws[f'B{row_num}'] = item.product.name
        ws[f'C{row_num}'] = item.quantity
        ws[f'D{row_num}'] = float(item.price)
        ws[f'E{row_num}'] = float(item.cost())
        row_num += 1

    ws[f'D{row_num}'] = "ИТОГО:"
    ws[f'D{row_num}'].font = openpyxl.styles.Font(bold=True)
    ws[f'E{row_num}'] = float(order.total_price)
    ws[f'E{row_num}'].font = openpyxl.styles.Font(bold=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    filename = f'order_{order.id}.xlsx'
    filepath = os.path.join(settings.MEDIA_ROOT, 'orders', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)

    return filepath


def send_order_email(user, order, excel_file):
    subject = f'Ваш заказ #{order.id} оформлен'

    context = {
        'user': user,
        'order': order,
    }
    message = render_to_string('shop/email/order_confirmation.txt', context)

    email = EmailMessage(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email]
    )

    with open(excel_file, 'rb') as f:
        email.attach(f'order_{order.id}.xlsx', f.read(),
                     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    email.send()

    if os.path.exists(excel_file):
        os.remove(excel_file)


@login_required
def order_success(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return render(request, 'shop/order_success.html', {'order': order})

@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(TrashElement, id=item_id, trash__user=request.user)
    product_name = cart_item.product.name
    cart_item.delete()
    messages.success(request, f'Товар "{product_name}" удален из корзины')
    return redirect('cart_view')


@login_required
def cart_view(request):
    trash, created = Trash.objects.get_or_create(user=request.user)

    cart_items = trash.items.select_related('product')
    total = sum(item.cost() for item in cart_items)

    context = {
        'cart_items': cart_items,
        'total': total
    }
    return render(request, 'shop/cart/cart_view.html', context)

def signup(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                print(f"User created: {user.username}")  # Для отладки
                login(request, user)
                return redirect('product_list')
            except Exception as e:
                print(f"Error creating user: {e}")  # Логируем ошибку
                # Добавьте сообщение об ошибке
                messages.error(request, f"Ошибка при создании пользователя: {e}")
        else:
            # Логируем ошибки формы
            print("Form errors:")
            for field, errors in form.errors.items():
                for error in errors:
                    print(f"- {field}: {error}")
    else:
        form = UserCreationForm()
    return render(request, 'shop/auth/signup.html', {'form': form})

@login_required
def profile(request):
    return render(request, 'shop/auth/profile.html')

def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('product_list')
    else:
        form = AuthenticationForm()
    return render(request, 'shop/auth/login.html', {'form': form})
def cart_count(request):
    if request.user.is_authenticated:
        count = TrashElement.objects.filter(trash__user=request.user).aggregate(
            total=Sum('quantity')
        )['total'] or 0
        return {'cart_count': count}
    return {'cart_count': 0}

def main(request):
    return HttpResponse('<div style="display: flex; align-items: center; flex-direction: column;"><h1>Добро пожаловать!</h1><div style="display: flex; justify-content: space-around; width: 100%"><h2 style="width: 50%"><a href="author/">Автор</a></h2><h2><a href="shop/">Магазин</a></h2></div></div>')